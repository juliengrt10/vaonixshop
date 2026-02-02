import openpyxl
import pandas as pd
import re
import math

# Configuration
INPUT_FILE = 'Liste de prix Vaonix 27022025.xlsm'
OUTPUT_FILE = 'vaonix_shopify_import_v2.csv'
PRICE_MULTIPLIER = 1.15
VENDOR_NAME = 'Vaonix'

# Missing "Classic" Products to inject
MISSING_PRODUCTS = [
    {
        'Bom': 'QSFP-DD-400G-FR4',
        'Description': '400G QSFP-DD FR4 1310nm 2km Duplex LC SMF Optical Transceiver Module',
        'Price': 650.00,  # Estimated base price
        'Category': 'QSFP-DD',
        'Tags': '400G, QSFP-DD, FR4, Transceiver'
    },
    {
        'Bom': 'QSFP-DD-400G-DR4',
        'Description': '400G QSFP-DD DR4 1310nm 500m MPO-12 SMF Optical Transceiver Module',
        'Price': 580.00,
        'Category': 'QSFP-DD',
        'Tags': '400G, QSFP-DD, DR4, Transceiver'
    },
    {
        'Bom': 'QSFP-DD-400G-SR8',
        'Description': '400G QSFP-DD SR8 850nm 100m MPO-16 MMF Optical Transceiver Module',
        'Price': 450.00,
        'Category': 'QSFP-DD',
        'Tags': '400G, QSFP-DD, SR8, Transceiver'
    },
    {
        'Bom': 'QSFP28-100G-LR4',
        'Description': '100G QSFP28 LR4 1310nm 10km LC SMF Transceiver Module',
        'Price': 180.00,
        'Category': 'QSFP28',
        'Tags': '100G, QSFP28, LR4, Transceiver'
    }
]

def clean_price(value):
    if value is None:
        return 0.0
    try:
        # Remove currency symbols or non-numeric chars if any
        if isinstance(value, str):
            value = re.sub(r'[^\d.,]', '', value).replace(',', '.')
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def parse_dac_length(bom):
    match = re.search(r'^(.*?)-(\d+(?:-\d+)?)M(.*)$', bom, re.IGNORECASE)
    if match:
        base = match.group(1)
        len_str = match.group(2)
        suffix = match.group(3)
        length_val = len_str.replace('-', '.')
        clean_suffix = suffix.strip('- ')
        
        variant_name = f"{length_val}m"
        if clean_suffix:
            variant_name += f" ({clean_suffix})"
            
        return base, variant_name, float(length_val)
    return None, None, None

def check_temp_variant(bom):
    # Detect -I (Industrial) or -E (Extended)
    match = re.search(r'^(.*?)(-(?:I|E))$', bom, re.IGNORECASE)
    if match:
        base = match.group(1)
        suffix = match.group(2).upper().strip('-')
        variant_name = "Industrial (-40/+85°C)" if suffix == 'I' else "Extended (-10/+80°C)"
        weight = 3 if suffix == 'I' else 2
        return base, variant_name, weight
    return None, None, None

def check_dwdm_variant(bom):
    # Detect DWDM Channel Cxx (e.g. -C17, -C61) anywhere in suffix
    # Assumes format contains -C17...
    match = re.search(r'^(.*?)(-C(\d{2}))(.*)$', bom, re.IGNORECASE)
    if match:
        base_part1 = match.group(1)
        channel_code = match.group(2) # -C17
        channel_num = match.group(3) # 17
        suffix_rest = match.group(4)
        
        base_bom = base_part1 + suffix_rest # Reconstruct base without channel? Or strictly group by Base-DWDM
        # Usually DWDM base is SFP-10G-ZR-DWDM
        # And variants are SFP-10G-ZR-DWDM-C17
        
        # Simpler regex: ends with -C\d{2}
        if re.search(r'-C\d{2}$', bom, re.IGNORECASE):
             base = re.sub(r'-C\d{2}$', '', bom, flags=re.IGNORECASE)
             channel = int(channel_num)
             # Map channel to Wavelength provided in helper if needed, or just use CX
             variant_name = f"Channel {channel} (C{channel})"
             return base, variant_name, channel
             
    return None, None, None

def check_cwdm_variant(bom):
    match = re.search(r'^(.*?)-(\d{4})(?:NM)?(-[IE])?$', bom, re.IGNORECASE)
    if match:
        base = match.group(1)
        wave = match.group(2)
        suffix = match.group(3)
        if 1270 <= int(wave) <= 1610:
            variant_name = f"{wave}nm"
            if suffix:
                 suffix_clean = suffix.replace('-', '').upper()
                 if suffix_clean == 'I': variant_name += " (Ind.)"
                 if suffix_clean == 'E': variant_name += " (Ext.)"
            return base, variant_name, int(wave)
    return None, None, None

def categorize_product(bom, description):
    bom_u = str(bom).upper()
    desc_u = str(description).upper()
    
    tags = []
    product_type = "Optical Transceiver"
    
    if 'DAC' in bom_u or 'CABLE' in desc_u:
        product_type = "Network Cable"
        tags.append("DAC")
        tags.append("Cable")
    elif 'AOC' in bom_u:
         product_type = "Active Optical Cable"
         tags.append("AOC")
         tags.append("Cable")
    else:
        tags.append("Transceiver")

    if 'QSFP-DD' in bom_u: tags.append('QSFP-DD')
    elif 'QSFP28' in bom_u: tags.append('QSFP28')
    elif 'QSFP+' in bom_u or 'QSFP' in bom_u: tags.append('QSFP+')
    elif 'SFP28' in bom_u: tags.append('SFP28')
    elif 'SFP+' in bom_u: tags.append('SFP+')
    elif 'SFP' in bom_u: tags.append('SFP')
    elif 'XFP' in bom_u: tags.append('XFP')
    
    if 'DWDM' in bom_u or 'DWDM' in desc_u: tags.append('DWDM')
    if 'CWDM' in bom_u or 'CWDM' in desc_u: tags.append('CWDM')
    if 'BIDI' in bom_u or 'BIDI' in desc_u: tags.append('BiDi')
    if 'TUNABLE' in desc_u: tags.append('Tunable')
    
    if '400G' in bom_u: tags.append('400G')
    elif '100G' in bom_u: tags.append('100G')
    elif '40G' in bom_u: tags.append('40G')
    elif '25G' in bom_u: tags.append('25G')
    elif '10G' in bom_u: tags.append('10G')
    elif '1G' in bom_u: tags.append('1G')

    return product_type, ", ".join(tags)

def main():
    print("Loading Excel file...")
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws = wb['Liste de prix']
    
    shopify_rows = []
    
    # Grouping dictionary
    # Key: Base Handle, Value: List of variants
    product_groups = {}
    
    # Standalone products
    standalone_products = []

    print("Processing rows...")
    start_row = 2 
    
    for row_idx in range(start_row, ws.max_row + 1):
        bom = ws.cell(row=row_idx, column=1).value
        description = ws.cell(row=row_idx, column=2).value
        if not bom or not description:
            continue

        if '-HP' in str(bom).upper() or ' HW' in str(bom).upper() or ' HW' in str(description).upper():
             continue
             
        price_raw = ws.cell(row=row_idx, column=9).value
        price_cost = clean_price(price_raw)
        final_price = round(price_cost * PRICE_MULTIPLIER, 2)
        
        product_type, tags = categorize_product(bom, description)
        
        # Base Item Data
        item = {
            'Handle': str(bom).lower().replace(' ', '-').replace('--', '-'),
            'Title': str(bom), 
            'Body (HTML)': description,
            'Vendor': VENDOR_NAME,
            'Type': product_type,
            'Tags': tags,
            'Published': 'TRUE',
            'Option1 Name': 'Title', 
            'Option1 Value': 'Default Title',
            'Variant Grams': 100, # Default weight
            'Variant Inventory Policy': 'deny', 
            'Variant Inventory Qty': 100, 
            'Variant Price': final_price,
            'Variant Compare At Price': '',
            'Image Src': '' 
        }
        
        # LOGIC 1: CABLES (DAC/AOC)
        is_cable = 'DAC' in str(bom).upper() or ('AOC' in str(bom).upper() and 'CABLE' in str(description).upper())
        
        # LOGIC 2: TEMPERATURE VARIANTS
        base_temp, var_name_temp, weight_temp = check_temp_variant(bom)
        
        # LOGIC 3: DWDM CHANNELS
        base_dwdm, var_name_dwdm, weight_dwdm = check_dwdm_variant(bom)
        
        # LOGIC 4: CWDM WAVELENGTHS
        base_cwdm, var_name_cwdm, weight_cwdm = check_cwdm_variant(bom)
        
        if is_cable:
            base_bom, variant_name, length_val = parse_dac_length(bom)
            if base_bom:
                group_key = base_bom.lower().replace(' ', '-')
                if group_key not in product_groups:
                    product_groups[group_key] = {'type': 'length', 'base_title': base_bom, 'variants': []}
                
                item['Option1 Name'] = 'Length'
                item['Option1 Value'] = variant_name
                item['_sort'] = length_val
                product_groups[group_key]['variants'].append(item)
            else:
                standalone_products.append(item)
                
        elif base_cwdm:
            group_key = base_cwdm.lower().replace(' ', '-')
            if group_key not in product_groups:
                product_groups[group_key] = {'type': 'wavelength', 'base_title': base_cwdm, 'variants': []}
            
            item['Option1 Name'] = 'Wavelength'
            item['Option1 Value'] = var_name_cwdm
            item['_sort'] = weight_cwdm
            product_groups[group_key]['variants'].append(item)

        elif base_dwdm:
            group_key = base_dwdm.lower().replace(' ', '-')
            if group_key not in product_groups:
                product_groups[group_key] = {'type': 'channel', 'base_title': base_dwdm, 'variants': []}
                
            item['Option1 Name'] = 'Channel (ITU)'
            item['Option1 Value'] = var_name_dwdm
            item['_sort'] = weight_dwdm
            product_groups[group_key]['variants'].append(item)
            
        elif base_temp:
            group_key = base_temp.lower().replace(' ', '-')
            # If we already have a commercial version (check below), we merge.
            # If checking order matters: usually Commercial comes first as base.
            if group_key not in product_groups:
                 product_groups[group_key] = {'type': 'temp', 'base_title': base_temp, 'variants': []}
            
            item['Option1 Name'] = 'Temperature'
            item['Option1 Value'] = var_name_temp
            item['_sort'] = weight_temp
            product_groups[group_key]['variants'].append(item)
            
        else:
            # Could be a Base for Temp or Channel, or purely standalone
            # We treat it as standalone first, but index it by Handle so we can merge later?
            # Simpler: All non-variants are added to 'standalone_products'
            # EXCEPT if they are the "Commercial" base of a Temperature group.
            
            # Check if this BOM is referenced as a base by others? 
            # It's hard in one pass. 
            # Strategy: Add to standalone. Post-process to merge standalone with groups if keys match.
            item['_is_base'] = True
            item['Option1 Name'] = 'Temperature' 
            item['Option1 Value'] = 'Commercial (0/70°C)'
            item['_sort'] = 1
            standalone_products.append(item)

    # MERGE STANDALONE INTO GROUPS
    final_standalone = []
    
    for item in standalone_products:
        handle = item['Handle']
        
        # Check if this handle exists as a group key
        if handle in product_groups:
            # It's the base product of a group! Add it as a variant.
            group_type = product_groups[handle]['type']
            
            if group_type == 'temp':
                 item['Option1 Name'] = 'Temperature'
                 item['Option1 Value'] = 'Commercial (0/70°C)'
            elif group_type == 'channel':
                 # Rare for DWDM base to exist as a sellable unit without channel, but if so:
                 item['Option1 Name'] = 'Channel (ITU)'
                 item['Option1 Value'] = 'Tunable / Unspecified' # Setup as needed
            elif group_type == 'wavelength':
                 item['Option1 Name'] = 'Wavelength'
                 item['Option1 Value'] = 'Unspecified'
            elif group_type == 'length':
                 # Base DAC usually not sellable without length, but just in case
                 item['Option1 Name'] = 'Length'
                 item['Option1 Value'] = 'Standard'
            
            product_groups[handle]['variants'].append(item)
        else:
            # Truly standalone (reset options to Default)
            item['Option1 Name'] = 'Title'
            item['Option1 Value'] = 'Default Title'
            final_standalone.append(item)

    # GENERATE ROWS FROM GROUPS
    for key, group in product_groups.items():
        variants = group['variants']
        variants.sort(key=lambda x: x['_sort'])
        
        for variant in variants:
            row = variant.copy()
            del row['_sort']
            if '_is_base' in row: del row['_is_base']
            
            row['Handle'] = key
            row['Title'] = group['base_title']
            shopify_rows.append(row)

    # GENERATE ROWS FROM STANDALONE
    for prod in final_standalone:
        row = prod.copy()
        if '_sort' in row: del row['_sort']
        if '_is_base' in row: del row['_is_base']
        shopify_rows.append(row)
        
    # Add Missing "Classic" Products
    for missing in MISSING_PRODUCTS:
        price = round(missing['Price'] * 1.15, 2)
        row = {
            'Handle': missing['Bom'].lower().replace(' ', '-'),
            'Title': missing['Bom'],
            'Body (HTML)': missing['Description'],
            'Vendor': VENDOR_NAME,
            'Type': 'Optical Transceiver',
            'Tags': missing['Tags'],
            'Published': 'TRUE',
            'Option1 Name': 'Title',
            'Option1 Value': 'Default Title',
            'Variant Grams': 100,
            'Variant Inventory Policy': 'deny',
            'Variant Inventory Qty': 50,
            'Variant Price': price,
            'Variant Compare At Price': '',
            'Image Src': ''
        }
        shopify_rows.append(row)

    # Convert to DataFrame
    df = pd.DataFrame(shopify_rows)
    
    # Define Column Order matching detailed Shopify Import
    columns = [
        'Handle', 'Title', 'Body (HTML)', 'Vendor', 'Type', 'Tags', 'Published',
        'Option1 Name', 'Option1 Value', 'Variant Grams', 'Variant Inventory Qty',
        'Variant Inventory Policy', 'Variant Price', 'Variant Compare At Price', 'Image Src'
    ]
    
    # Reindex checks if columns exist, fills missing with Nan
    df = df.reindex(columns=columns)
    
    print(f"Generated {len(df)} rows.")
    df.to_csv(OUTPUT_FILE, index=False)
    print(f"Saved to {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
