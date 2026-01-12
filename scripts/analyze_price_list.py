import openpyxl
import pandas as pd

# Load the Excel file
wb = openpyxl.load_workbook('Liste de prix Vaonix 27022025.xlsm', data_only=True)

# Analyze the main sheet
ws = wb['Liste de prix']

print("=" * 80)
print("ANALYSIS OF 'Liste de prix' SHEET")
print("=" * 80)
print(f"\nDimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")

# Print first 20 rows to understand structure
print("\n" + "=" * 80)
print("FIRST 20 ROWS:")
print("=" * 80)

for row_idx in range(1, min(21, ws.max_row + 1)):
    row_data = []
    for col_idx in range(1, min(15, ws.max_column + 1)):  # Limit to first 15 columns
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value
        if value is not None:
            row_data.append(f"Col{col_idx}: {str(value)[:50]}")
    
    if row_data:
        print(f"\nRow {row_idx}:")
        for item in row_data:
            print(f"  {item}")

# Try to find header row
print("\n" + "=" * 80)
print("LOOKING FOR HEADERS (rows containing 'BOM', 'Prix', 'Description'):")
print("=" * 80)

for row_idx in range(1, min(50, ws.max_row + 1)):
    row_values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]
    row_str = ' | '.join([str(v) if v else '' for v in row_values[:15]])
    
    if any(keyword in str(row_str).upper() for keyword in ['BOM', 'PRIX', 'DESCRIPTION', 'REFERENCE', 'DÉSIGNATION']):
        print(f"\nRow {row_idx}: {row_str}")

wb.close()
